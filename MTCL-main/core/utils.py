'''
* @name: utils.py
* @description: Other functions.
'''


import os
import random
import numpy as np
import torch
import torch.nn as nn
from sklearn.metrics import accuracy_score, f1_score,confusion_matrix
from torch.nn.utils.rnn import pack_padded_sequence,pad_packed_sequence
from openpyxl import Workbook,load_workbook
import torch.nn.functional as F

def Loss_Branch(R_f, R_c, label_u, label_m,loss_js, eliminate_rate):
    loss_fn = torch.nn.MSELoss()
    loss_fn_none = torch.nn.MSELoss(reduction='none')

    loss_c = loss_fn(R_c,label_u)
    loss_js,small_loss_idx,big_loss_idx = eliminated_confilict(loss_js, eliminate_rate)
    loss_f = loss_fn_none(R_f,label_u)
    loss_change = loss_fn_none(R_f,label_m)

    loss_sum = loss_c + loss_js + torch.mean(loss_f[small_loss_idx]) + torch.mean(loss_change[big_loss_idx])
    return loss_sum

def JS_divergence(feat_1, feat_2):
    M = 0.5 * (F.softmax(feat_1, dim=1) + F.softmax(feat_2, dim=1))
    kl_1 = F.kl_div(F.log_softmax(feat_1, dim=1), M, reduction='none').sum(dim=1)
    kl_2 = F.kl_div(F.log_softmax(feat_2, dim=1), M, reduction='none').sum(dim=1)
    js_div = 0.5 * (kl_1 + kl_2)
    return js_div / torch.log(torch.tensor(2.0))

def eliminated_confilict(loss_js, eliminate_rate):

    conflict_loss = loss_js.cpu()
    sorted_idx = np.argsort(conflict_loss.data)
    loss_sorted = conflict_loss[sorted_idx]
    update_idx = sorted_idx[:int((1 - eliminate_rate) * len(loss_sorted))]
    bigger_idx = sorted_idx[int((1 - eliminate_rate) * len(loss_sorted)):]
    loss_eliminated = torch.mean(conflict_loss[update_idx])
    return loss_eliminated,update_idx,bigger_idx

    
class AverageMeter(object):
    def __init__(self):
        self.value = 0
        self.value_avg = 0
        self.value_sum = 0
        self.count = 0

    def reset(self):
        self.value = 0
        self.value_avg = 0
        self.value_sum = 0
        self.count = 0

    def update(self, value, count):
        self.value = value
        self.value_sum += value * count
        self.count += count
        self.value_avg = self.value_sum / self.count


def setup_seed(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def save_model(save_path, epoch, model, optimizer):
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    save_file_path = os.path.join(save_path,'best.pth')
    states = {
        'epoch': epoch + 1,
        'state_dict': model.state_dict(),
        'optimizer': optimizer.state_dict(),
    }
    torch.save(states, save_file_path)
    
def save_args(opt, file_path):
    with open(file_path, "w") as file:
        for key, value in vars(opt).items():
            file.write(f"{key}: {value}\n")

def save_excel(name, file_path, best_result,best_epoch):
    if os.path.isfile(file_path):
        workbook_all = load_workbook(file_path)
        sheet = workbook_all['Sheet']
        sheet.append([name,str(best_result['Mult_acc_2']),str(best_result['Mult_acc_3']),str(best_result['Mult_acc_5']),str(best_result['F1_score']),str(best_result['MAE']),str(best_result['Corr']),best_epoch])
        workbook_all.save(file_path)
    else:
        workbook_all = Workbook()
        sheet = workbook_all.active
        sheet.append([f'model', 'Mult_acc_2', 'Mult_acc_3', 'Mult_acc_5', 'F1_score', 'MAE','Corr','epoch'])
        sheet.append([name,str(best_result['Mult_acc_2']),str(best_result['Mult_acc_3']),str(best_result['Mult_acc_5']),str(best_result['F1_score']),str(best_result['MAE']),str(best_result['Corr']),best_epoch])
        workbook_all.save(file_path)
